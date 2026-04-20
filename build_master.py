"""
Build the Master xlsx file for Prandos Dashboard.
Contains all telemarketing (Excel) and social (PDF) data in normalized tabs.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
import json
from pathlib import Path

OUT_DIR = Path("/Users/zecharialandau/Zecharia-Cloude/the-system-v8/dashboard-prandos")

# ============================================================
# NORMALIZED DATA
# ============================================================

# Telemarketing - from Prandos Excel files
telemarketing = [
    {
        "month_key": "2025-12", "month_label": "דצמבר 2025",
        "deals": 37021, "billings": 27196,
        "target_pct": 0.8776, "target_deals_pct": None, "target_billings_pct": None,
        "agent_hours": 91.7, "agent_cost": 7794.5,
        "manager_hours": 117.7, "manager_cost": 15654.1,
        "commissions": 5553.15, "total_expenses": 29001.75,
        "campaign_revenue": 40278,
        "prandos_commission": 10298.64,
        "prandos_commission_base": 85822,
        "pay_peach_current": 26284, "pay_peach_recurring": 26284,
        "pay_cardcom": 50, "pay_paybox": 612, "pay_bank": 250, "pay_jgive": 0,
    },
    {
        "month_key": "2026-01", "month_label": "ינואר 2026",
        "deals": 169413, "billings": 112603,
        "target_pct": 0.996, "target_deals_pct": 1.0912, "target_billings_pct": 8.7032,
        "agent_hours": 369.74, "agent_cost": 31427.9,
        "manager_hours": 137.1, "manager_cost": 18234.3,
        "commissions": 25412, "total_expenses": 75074.2,
        "campaign_revenue": 3577,
        "prandos_commission": 13422.18,
        "prandos_commission_base": 111851.5,
        "pay_peach_current": 82721, "pay_peach_recurring": 87050,
        "pay_cardcom": 1576, "pay_paybox": 12900, "pay_bank": 11077, "pay_jgive": 0,
    },
    {
        "month_key": "2026-02", "month_label": "פברואר 2026",
        "deals": 132005, "billings": 93182,
        "target_pct": 0.5963, "target_deals_pct": 0.85027, "target_billings_pct": 3.6012,
        "agent_hours": 480.18, "agent_cost": 40815.3,
        "manager_hours": 181.43, "manager_cost": 21000,
        "commissions": 19800, "total_expenses": 81615.3,
        "campaign_revenue": 17383,
        "prandos_commission": 11916,
        "prandos_commission_base": 99303,
        "pay_peach_current": 80487, "pay_peach_recurring": 83403,
        "pay_cardcom": 3054, "pay_paybox": 3855, "pay_bank": 2870, "pay_jgive": 0,
    },
    {
        "month_key": "2026-03", "month_label": "מרץ 2026",
        "deals": 129767, "billings": 121541,
        "target_pct": 0.8171, "target_deals_pct": 0.8358, "target_billings_pct": 3.1314,
        "agent_hours": 345.24, "agent_cost": 29345.4,
        "manager_hours": 172.48, "manager_cost": 21000,
        "commissions": 19465.05, "total_expenses": 69810.45,
        "campaign_revenue": 29794,
        "prandos_commission": 16026.36,
        "prandos_commission_base": 133553,
        "pay_peach_current": 81137, "pay_peach_recurring": 98919,
        "pay_cardcom": 9474, "pay_paybox": 6238, "pay_bank": 6160, "pay_jgive": 750,
    },
]

# Cumulative since opening (from new master file)
cumulative = {
    "total_deals_since_opening": 468206,
    "total_billings_since_opening": 354522,
    "soldiers_adopted": 56,
}

# Social: Facebook, Instagram, TikTok
social_facebook = [
    {"month_key": "2025-Q4", "month_label": "Q4 2025 (אוק-דצמ)",
     "views": 5677906, "views_3sec": 1500000, "views_1min": 12200,
     "content_interactions": 186000, "visits": 28100, "new_follows": 2700,
     "total_followers": 24610, "views_organic": 791182, "views_ads": 4886724,
     "viewers": 1453247, "watch_time_days": None},
    {"month_key": "2026-01", "month_label": "ינואר 2026",
     "views": 1681526, "views_3sec": 463800, "views_1min": 340,
     "content_interactions": 61000, "visits": 7700, "new_follows": 1800,
     "total_followers": 25202, "views_organic": 323565, "views_ads": 1357961,
     "viewers": 772111, "watch_time_days": 81.375},
    {"month_key": "2026-02", "month_label": "פברואר 2026",
     "views": 1826677, "views_3sec": 409800, "views_1min": 3000,
     "content_interactions": 35700, "visits": 9100, "new_follows": 726,
     "total_followers": 26418, "views_organic": 172442, "views_ads": 1654235,
     "viewers": 809068, "watch_time_days": 70.29},
    {"month_key": "2026-03", "month_label": "מרץ 2026",
     "views": 4186348, "views_3sec": 815000, "views_1min": 29200,
     "content_interactions": 121500, "visits": 14400, "new_follows": 6000,
     "total_followers": 31541, "views_organic": 835962, "views_ads": 3350386,
     "viewers": 1565644, "watch_time_days": 182.79},
]

social_instagram = [
    {"month_key": "2025-Q4", "month_label": "Q4 2025 (אוק-דצמ)",
     "views": 2307404, "reach": 716900, "content_interactions": 66000,
     "views_organic": 218324, "views_ads": 2089080,
     "followers": 7029, "growth": 571, "follows": 638, "unfollows": 67,
     "accounts_reached": 634650},
    {"month_key": "2026-01", "month_label": "ינואר 2026",
     "views": 636525, "reach": 276200, "content_interactions": 17900,
     "views_organic": 43211, "views_ads": 593314,
     "followers": 7091, "growth": 91, "follows": 109, "unfollows": 18,
     "accounts_reached": 276156},
    {"month_key": "2026-02", "month_label": "פברואר 2026",
     "views": 803804, "reach": 423300, "content_interactions": 17000,
     "views_organic": 112341, "views_ads": 691463,
     "followers": 7419, "growth": 178, "follows": 208, "unfollows": 30,
     "accounts_reached": 276156},
    {"month_key": "2026-03", "month_label": "מרץ 2026",
     "views": 1339667, "reach": 615600, "content_interactions": 34800,
     "views_organic": 193570, "views_ads": 1146097,
     "followers": 7669, "growth": 399, "follows": 454, "unfollows": 55,
     "accounts_reached": 615584},
]

social_tiktok = [
    {"month_key": "2025-Q4", "month_label": "Q4 2025 (אוק-דצמ)",
     "video_views": 5500000, "profile_views": 21000, "likes": 79000,
     "comments": 1800, "shares": 3600,
     "net_follows": 3200, "total_followers": 12000},
    {"month_key": "2026-01", "month_label": "ינואר 2026",
     "video_views": 810000, "profile_views": 3700, "likes": 11000,
     "comments": 288, "shares": 385,
     "net_follows": 235, "total_followers": 12000},
    {"month_key": "2026-02", "month_label": "פברואר 2026",
     "video_views": 841000, "profile_views": 3900, "likes": 10000,
     "comments": 232, "shares": 561,
     "net_follows": 275, "total_followers": 12000},
    {"month_key": "2026-03", "month_label": "מרץ 2026",
     "video_views": 1500000, "profile_views": 5300, "likes": 20000,
     "comments": 327, "shares": 746,
     "net_follows": 479, "total_followers": 13000},
]

# ============================================================
# BUILD THE WORKBOOK
# ============================================================

wb = openpyxl.Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
thin = Side(border_style="thin", color="999999")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_table(ws, headers, rows, start_row=1):
    ws.sheet_view.rightToLeft = True
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.alignment = Alignment(horizontal="center")
            c.border = BORDER
            if isinstance(val, (int, float)) and c_idx > 2:
                c.number_format = '#,##0.##'
    # autosize
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(str(h)) + 3)

# --- Tab 1: טלמרקטינג ---
ws = wb.create_sheet("telemarketing")
headers = ["חודש", "תווית", "עיסקאות", "סליקות", "יעד %", "שעות נציגים",
           "עלות נציגים", "שעות מנהל", "עלות מנהל", "עמלות נציגים",
           "סה\"כ הוצאות מוקד", "הכנסות קמפיינים", "בסיס עמלה פרנדוס", "עמלת פרנדוס 12%"]
rows = [[r["month_key"], r["month_label"], r["deals"], r["billings"], r["target_pct"],
         r["agent_hours"], r["agent_cost"], r["manager_hours"], r["manager_cost"],
         r["commissions"], r["total_expenses"], r["campaign_revenue"],
         r["prandos_commission_base"], r["prandos_commission"]]
        for r in telemarketing]
write_table(ws, headers, rows)

# --- Tab 2: Facebook ---
ws = wb.create_sheet("facebook")
headers = ["חודש", "תווית", "Views", "3-sec views", "1-min views",
           "Content interactions", "Visits", "עוקבים חדשים", "סה\"כ עוקבים",
           "Views אורגני", "Views ממומן", "Viewers", "Watch time (ימים)"]
rows = [[r["month_key"], r["month_label"], r["views"], r["views_3sec"], r["views_1min"],
         r["content_interactions"], r["visits"], r["new_follows"], r["total_followers"],
         r["views_organic"], r["views_ads"], r["viewers"], r["watch_time_days"]]
        for r in social_facebook]
write_table(ws, headers, rows)

# --- Tab 3: Instagram ---
ws = wb.create_sheet("instagram")
headers = ["חודש", "תווית", "Views", "Reach", "Content interactions",
           "Views אורגני", "Views ממומן", "עוקבים", "גידול נטו",
           "Follows", "Unfollows", "Accounts Reached"]
rows = [[r["month_key"], r["month_label"], r["views"], r["reach"], r["content_interactions"],
         r["views_organic"], r["views_ads"], r["followers"], r["growth"],
         r["follows"], r["unfollows"], r["accounts_reached"]]
        for r in social_instagram]
write_table(ws, headers, rows)

# --- Tab 4: TikTok ---
ws = wb.create_sheet("tiktok")
headers = ["חודש", "תווית", "Video Views", "Profile Views", "Likes",
           "Comments", "Shares", "גידול נטו עוקבים", "סה\"כ עוקבים"]
rows = [[r["month_key"], r["month_label"], r["video_views"], r["profile_views"], r["likes"],
         r["comments"], r["shares"], r["net_follows"], r["total_followers"]]
        for r in social_tiktok]
write_table(ws, headers, rows)

# --- Tab 5: סיכום חודשי ---
ws = wb.create_sheet("monthly_summary")
ws.sheet_view.rightToLeft = True
headers = ["חודש", "תווית",
           "עיסקאות מוקד", "סליקות מוקד", "רווח מוקד נטו",
           "FB Views", "IG Views", "TT Views", "סה\"כ Views סושיאל",
           "FB Interactions", "IG Interactions", "TT Likes+Comments",
           "סה\"כ Followers (FB+IG+TT)", "עמלת פרנדוס"]
rows = []
for i, tm in enumerate(telemarketing):
    # find matching social month
    fb = next((s for s in social_facebook if s["month_key"] == tm["month_key"]), None)
    ig = next((s for s in social_instagram if s["month_key"] == tm["month_key"]), None)
    tt = next((s for s in social_tiktok if s["month_key"] == tm["month_key"]), None)
    net_profit = (tm["billings"] + tm["campaign_revenue"]) - tm["total_expenses"] - tm["prandos_commission"]
    fb_v = fb["views"] if fb else 0
    ig_v = ig["views"] if ig else 0
    tt_v = tt["video_views"] if tt else 0
    fb_i = fb["content_interactions"] if fb else 0
    ig_i = ig["content_interactions"] if ig else 0
    tt_i = (tt["likes"] + tt["comments"]) if tt else 0
    followers_total = (fb["total_followers"] if fb else 0) + (ig["followers"] if ig else 0) + (tt["total_followers"] if tt else 0)
    rows.append([tm["month_key"], tm["month_label"],
                 tm["deals"], tm["billings"], net_profit,
                 fb_v, ig_v, tt_v, fb_v + ig_v + tt_v,
                 fb_i, ig_i, tt_i,
                 followers_total, tm["prandos_commission"]])
write_table(ws, headers, rows)

# reorder so monthly_summary is first
wb.move_sheet("monthly_summary", offset=-4)

# --- Tab 6: README ---
ws = wb.create_sheet("README")
ws.sheet_view.rightToLeft = True
ws["A1"] = "מאגר נתונים - דאשבורד פרנדוס + גדולים במדים"
ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
ws.merge_cells("A1:F1")

content = [
    "",
    "מבנה הקובץ:",
    "• monthly_summary — סיכום חודשי משוקלל (כל ה-KPIs במקום אחד)",
    "• telemarketing — נתוני מוקד פרנדוס לפי חודש",
    "• facebook — מטריקות פייסבוק לפי חודש",
    "• instagram — מטריקות אינסטגרם לפי חודש",
    "• tiktok — מטריקות טיקטוק לפי חודש",
    "",
    "איך מוסיפים חודש חדש?",
    "1. שלח את קובץ ה-Excel של פרנדוס ואת ה-PDF של גדולים במדים לצ'אט",
    "2. אני מוסיף שורה חדשה בכל Tab רלוונטי",
    "3. הדאשבורד (dashboard.html) מתעדכן אוטומטית",
    "",
    "הערה: נתוני Q4 2025 (אוקטובר-דצמבר) מגיעים מדוח רבעוני אחד, ולכן מוצגים יחד.",
    "החל מינואר 2026 הנתונים חודשיים.",
]
for i, txt in enumerate(content, 2):
    c = ws.cell(row=i, column=1, value=txt)
    c.alignment = Alignment(horizontal="right", vertical="center")
    if txt.startswith(("מבנה", "איך")):
        c.font = Font(bold=True, size=12, color="1F4E78")
ws.column_dimensions["A"].width = 80

wb.move_sheet("README", offset=-5)

OUT_XLSX = OUT_DIR / "Master-Dashboard-Data.xlsx"
wb.save(OUT_XLSX)

# Also export to JSON so the HTML dashboard can load it without internet
data_for_js = {
    "telemarketing": telemarketing,
    "facebook": social_facebook,
    "instagram": social_instagram,
    "tiktok": social_tiktok,
    "cumulative": cumulative,
    "last_updated": "2026-04-20",
}
(OUT_DIR / "data.json").write_text(json.dumps(data_for_js, ensure_ascii=False, indent=2), encoding="utf-8")

print(f"✓ Saved: {OUT_XLSX}")
print(f"✓ Saved: {OUT_DIR / 'data.json'}")
print(f"Months: {[t['month_label'] for t in telemarketing]}")
